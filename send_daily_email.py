"""
Daily Email Sender — SAP Job Listings
Uses Microsoft Graph API (Azure app permissions — no SMTP needed).
Triggered by GitHub Actions (see .github/workflows/*.yml) — run directly with no arguments.

Required secrets / .env:
    SUPABASE_URL, SUPABASE_KEY
    AZURE_TENANT_ID, AZURE_CLIENT_ID, AZURE_CLIENT_SECRET
    EMAIL_FROM   — the mailbox to send from  (e.g. automation@yourcompany.com)
    EMAIL_TO     — comma-separated recipients
    EMAIL_CC     — comma-separated CC (optional)
"""

import os
import json
import io
import base64
import logging
import time
from datetime import datetime, date, timedelta

import requests
from dotenv import load_dotenv
from supabase import create_client

# ─────────────────────────────────────────────
# ENV
# ─────────────────────────────────────────────
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

SUPABASE_URL        = os.getenv("SUPABASE_URL")
SUPABASE_KEY        = os.getenv("SUPABASE_KEY")
AZURE_TENANT_ID     = os.getenv("AZURE_TENANT_ID")
AZURE_CLIENT_ID     = os.getenv("AZURE_CLIENT_ID")
AZURE_CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")
EMAIL_FROM          = os.getenv("EMAIL_FROM")
EMAIL_TO            = os.getenv("EMAIL_TO", "")
EMAIL_CC            = os.getenv("EMAIL_CC", "")
REPORT_LOOKBACK_HOURS = int(os.getenv("REPORT_LOOKBACK_HOURS", "24"))
NOTIFICATION_STATE_TABLE = os.getenv("NOTIFICATION_STATE_TABLE", "automation_state")
NOTIFICATION_STATE_KEY = os.getenv("NOTIFICATION_STATE_KEY", "sap_daily_email_last_sent_at")

for _var, _val in {
    "SUPABASE_URL": SUPABASE_URL, "SUPABASE_KEY": SUPABASE_KEY,
    "AZURE_TENANT_ID": AZURE_TENANT_ID, "AZURE_CLIENT_ID": AZURE_CLIENT_ID,
    "AZURE_CLIENT_SECRET": AZURE_CLIENT_SECRET,
    "EMAIL_FROM": EMAIL_FROM, "EMAIL_TO": EMAIL_TO,
}.items():
    if not _val:
        raise EnvironmentError(f"Missing required env var: {_var}")

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# ─────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)

# ─────────────────────────────────────────────
# GRAPH AUTH
# ─────────────────────────────────────────────
_token_cache = {"token": None, "expires_at": 0}


def get_graph_token() -> str:
    """Obtain (or return cached) Microsoft Graph access token via client credentials."""
    now = time.time()
    if _token_cache["token"] and now < _token_cache["expires_at"] - 60:
        return _token_cache["token"]

    resp = requests.post(
        f"https://login.microsoftonline.com/{AZURE_TENANT_ID}/oauth2/v2.0/token",
        data={
            "grant_type":    "client_credentials",
            "client_id":     AZURE_CLIENT_ID,
            "client_secret": AZURE_CLIENT_SECRET,
            "scope":         "https://graph.microsoft.com/.default",
        },
        timeout=20,
    )
    resp.raise_for_status()
    data = resp.json()
    _token_cache["token"]      = data["access_token"]
    _token_cache["expires_at"] = now + data.get("expires_in", 3600)
    logging.info("Graph access token acquired.")
    return _token_cache["token"]


# ─────────────────────────────────────────────
# DATA HELPERS
# ─────────────────────────────────────────────
COLUMNS_FROM_DB = [
    "jr_no",
    "skill_name",
    "posting_start_date",
    "posting_end_date",
    "client_recruiter",
    "recruiter_email",
    "job_details",
    "company_name",
    "jr_status",
]

DISPLAY_HEADERS = {
    "jr_no":               "JR No",
    "skill_name":          "Job Title",
    "posting_start_date":  "Posting Start Date",
    "posting_end_date":    "Posting End Date",
    "client_recruiter":    "Client Recruiter",
    "recruiter_email":     "Recruiter Email",
    "job_details":         "Job Details",
    "company_name":        "Company",
    "jr_status":           "JR Status",
}


def fetch_active_jobs() -> list:
    """Pull active records with posting_end_date >= today."""
    today_iso = date.today().isoformat()
    resp = (
        supabase.table("jr_master")
        .select(", ".join(COLUMNS_FROM_DB))
        .gte("posting_end_date", today_iso)
        .neq("jr_status", "inactive")
        .order("posting_start_date", desc=True)
        .limit(5000)
        .execute()
    )
    return resp.data or []


HANDOFF_FILE = "scraper_handoff.json"


def _fallback_report_start() -> datetime:
    return datetime.now() - timedelta(hours=REPORT_LOOKBACK_HOURS)


def get_last_successful_email_at() -> datetime:
    """Return the last successful email checkpoint, or a 24h fallback."""
    fallback = _fallback_report_start()
    try:
        resp = (
            supabase.table(NOTIFICATION_STATE_TABLE)
            .select("value")
            .eq("key", NOTIFICATION_STATE_KEY)
            .limit(1)
            .execute()
        )
        rows = resp.data or []
        if not rows or not rows[0].get("value"):
            logging.info(
                f"No notification checkpoint found; using {REPORT_LOOKBACK_HOURS}h fallback"
            )
            return fallback
        return datetime.fromisoformat(str(rows[0]["value"]))
    except Exception as e:
        logging.warning(
            f"Could not read notification checkpoint from {NOTIFICATION_STATE_TABLE}: {e}; "
            f"using {REPORT_LOOKBACK_HOURS}h fallback"
        )
        return fallback


def save_successful_email_checkpoint(sent_at: datetime) -> None:
    """Persist the report window end after a successful email send."""
    sent_at_iso = sent_at.isoformat()
    payload = {
        "key": NOTIFICATION_STATE_KEY,
        "value": sent_at_iso,
        "updated_at": datetime.now().isoformat(),
    }
    try:
        supabase.table(NOTIFICATION_STATE_TABLE).upsert(
            payload,
            on_conflict="key",
            ignore_duplicates=False,
        ).execute()
        logging.info(f"Notification checkpoint saved: {sent_at_iso}")
    except Exception as e:
        logging.error(
            f"Could not save notification checkpoint to {NOTIFICATION_STATE_TABLE}: {e}"
        )
        raise


def clear_legacy_new_jr_status() -> None:
    """Clear old status-based new-JR flags left by previous deployments."""
    now_iso = datetime.now().isoformat()
    try:
        resp = supabase.table("jr_master").select("jr_no").eq("jr_status", "new jr").execute()
        jr_nos = [r["jr_no"] for r in (resp.data or [])]
        if not jr_nos:
            return

        batch_size = 50
        for i in range(0, len(jr_nos), batch_size):
            batch = jr_nos[i: i + batch_size]
            supabase.table("jr_master").update(
                {"jr_status": "active", "modified_date": now_iso}
            ).in_("jr_no", batch).execute()
        logging.info(f"Cleared {len(jr_nos)} legacy new-jr status flag(s)")
    except Exception as e:
        logging.warning(f"Could not clear legacy new-jr status flags: {e}")


def _load_handoff() -> dict:
    """Read the JSON file written by the scraper. Returns empty lists if missing."""
    try:
        with open(HANDOFF_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        logging.warning(f"Handoff file '{HANDOFF_FILE}' not found — highlights will be empty")
        return {"new_jr_nos": [], "deactivated_jr_nos": []}
    except Exception as e:
        logging.warning(f"Could not read handoff file: {e}")
        return {"new_jr_nos": [], "deactivated_jr_nos": []}


def _fetch_rows_by_jr_nos(jr_nos: list) -> list:
    """Fetch jr_no, skill_name, posting_start_date, client_recruiter, jr_status
    for a given list of jr_nos. posting_start_date is needed for the Assign Date column.
    """
    if not jr_nos:
        return []
    cols = "jr_no, skill_name, posting_start_date, client_recruiter, jr_status"
    try:
        resp = (
            supabase.table("jr_master")
            .select(cols)
            .in_("jr_no", jr_nos)
            .order("jr_no")
            .execute()
        )
        return resp.data or []
    except Exception as e:
        logging.warning(f"_fetch_rows_by_jr_nos failed: {e}")
        return []


def _dedupe_rows(rows: list) -> list:
    """Return one row per jr_no while preserving first-seen order."""
    seen = set()
    deduped = []
    for row in rows:
        jr_no = row.get("jr_no")
        if not jr_no or jr_no in seen:
            continue
        seen.add(jr_no)
        deduped.append(row)
    return deduped


def _fetch_recent_changes(window_start: datetime, window_end: datetime) -> dict:
    """Fetch accumulated changes detected by hourly scraper runs."""
    start_iso = window_start.isoformat()
    end_iso = window_end.isoformat()
    cols = "jr_no, skill_name, posting_start_date, client_recruiter, jr_status"

    try:
        new_resp = (
            supabase.table("jr_master")
            .select(cols)
            .gte("created_date", start_iso)
            .lte("created_date", end_iso)
            .order("jr_no")
            .execute()
        )
        new_rows = new_resp.data or []
    except Exception as e:
        logging.warning(f"Recent new-jr fetch failed: {e}")
        new_rows = []

    try:
        legacy_new_resp = (
            supabase.table("jr_master")
            .select(cols)
            .eq("jr_status", "new jr")
            .order("jr_no")
            .execute()
        )
        legacy_new_rows = legacy_new_resp.data or []
    except Exception as e:
        logging.warning(f"Legacy new-jr fetch failed: {e}")
        legacy_new_rows = []

    try:
        deact_resp = (
            supabase.table("jr_master")
            .select(cols)
            .eq("jr_status", "inactive")
            .gte("modified_date", start_iso)
            .lte("modified_date", end_iso)
            .order("jr_no")
            .execute()
        )
        deact_rows = deact_resp.data or []
    except Exception as e:
        logging.warning(f"Recent deactivated fetch failed: {e}")
        deact_rows = []

    logging.info(
        f"Recent DB changes ({start_iso} to {end_iso}): "
        f"{len(new_rows)} created-date new jr, "
        f"{len(legacy_new_rows)} legacy new jr, "
        f"{len(deact_rows)} deactivated"
    )
    return {"new_jr": _dedupe_rows(new_rows + legacy_new_rows), "deactivated": deact_rows}


def fetch_highlights(window_start: datetime, window_end: datetime) -> dict:
    """Fetch accumulated hourly scraper changes for the daily notification.

    The handoff file is still merged in for same-run/manual executions, but the
    DB is the source that survives separate hourly GitHub Actions runs.
    """
    recent = _fetch_recent_changes(window_start, window_end)
    handoff = _load_handoff()
    handoff_new = _fetch_rows_by_jr_nos(handoff.get("new_jr_nos", []))
    handoff_deact = _fetch_rows_by_jr_nos(handoff.get("deactivated_jr_nos", []))

    return {
        "new_jr": _dedupe_rows(recent["new_jr"] + handoff_new),
        "deactivated": _dedupe_rows(recent["deactivated"] + handoff_deact),
    }


def fetch_summary_counts(highlights: dict) -> dict:
    """Derive deactivated count from highlights; query DB for active + new jr counts."""
    try:
        active_resp  = supabase.table("jr_master").select("jr_no").eq("jr_status", "active").execute()
        active_count = len(active_resp.data or [])
    except Exception as e:
        logging.warning(f"fetch_summary active count failed: {e}")
        active_count = 0

    new_count = len(highlights.get("new_jr", []))
    deact_count = len(highlights.get("deactivated", []))

    return {"active": active_count, "new_jr": new_count, "deactivated": deact_count}


# ─────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────
STATUS_COLORS = {
    "new jr":   "00B050",   # green
    "active":   "0070C0",   # blue
    "inactive": "FF0000",   # red
}


def build_excel(records: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Listings"

    thin   = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header ──
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_key in enumerate(COLUMNS_FROM_DB, start=1):
        cell = ws.cell(row=1, column=col_idx, value=DISPLAY_HEADERS.get(col_key, col_key))
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
    ws.row_dimensions[1].height = 28

    # ── Data ──
    alt_a = PatternFill("solid", fgColor="EBF3FB")
    alt_b = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, record in enumerate(records, start=2):
        row_fill   = alt_a if row_idx % 2 == 0 else alt_b
        status_raw = str(record.get("jr_status", "")).strip().lower()

        for col_idx, col_key in enumerate(COLUMNS_FROM_DB, start=1):
            value = record.get(col_key, "")
            if value is None or str(value) == "None":
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_key == "job_details"))
            cell.border    = border

            if col_key == "jr_status":
                cell.fill      = PatternFill("solid", fgColor=STATUS_COLORS.get(status_raw, "808080"))
                cell.font      = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = row_fill

    col_widths = {
        "jr_no": 12, "skill_name": 30, "posting_start_date": 18,
        "posting_end_date": 18, "client_recruiter": 22,
        "recruiter_email": 28, "job_details": 50,
        "company_name": 14, "jr_status": 14,
    }
    for col_idx, col_key in enumerate(COLUMNS_FROM_DB, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_key, 16)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
# HTML BODY HELPERS
# ─────────────────────────────────────────────

# Compact badge labels shown in the email table
# "D" = Deactivated Today  |  "A" = Active  |  "New" = New JR
_STATUS_BADGE = {
    "new jr":   ("New JR",         "#00B050"),
    "active":   ("Active",         "#0070C0"),
    "inactive": ("D",              "#C00000"),  # compact — "D" = Deactivated Today
}


def _fmt_date(val) -> str:
    """Format an ISO date string (or None) to DD-Mon-YY. Returns '' if blank."""
    if not val:
        return ""
    try:
        return datetime.fromisoformat(str(val)).strftime("%d-%b-%y")
    except Exception:
        return str(val)


def _highlights_table_html(rows: list, status: str) -> str:
    """Build a compact HTML highlight table for new jr or deactivated rows.

    Columns: JR No | Job Title | JR Assign Date | Recruiter | St
    ─────────────────────────────────────────────────────────────
    Changes vs original:
      • Added 'JR Assign Date' column (posting_start_date, DD-Mon-YY format)
      • Status badge shrunk to 'D' / 'A' / 'New' instead of full label text
      • Header & badge colour driven by _STATUS_BADGE lookup
    """
    if not rows:
        return ""

    badge_label, hdr_color = _STATUS_BADGE.get(status, ("?", "#888888"))
    section_title = "New JR" if status == "new jr" else "Deactivated Today"
    row_bg        = "#f0fff4" if status == "new jr" else "#fff5f5"

    rows_html = ""
    for i, r in enumerate(rows):
        bg          = row_bg if i % 2 == 0 else "#ffffff"
        recruiter   = r.get("client_recruiter") or "—"
        assign_date = _fmt_date(r.get("posting_start_date"))

        rows_html += (
            f"<tr style='background:{bg};'>"
            # JR No
            f"<td style='padding:5px 8px;border-bottom:1px solid #e8e8e8;"
            f"font-size:12px;white-space:nowrap;'>{r.get('jr_no', '')}</td>"
            # Job Title — truncate long strings via title tooltip
            f"<td style='padding:5px 8px;border-bottom:1px solid #e8e8e8;"
            f"font-size:12px;max-width:160px;overflow:hidden;"
            f"text-overflow:ellipsis;white-space:nowrap;' "
            f"title='{r.get('skill_name', '')}'>{r.get('skill_name', '')}</td>"
            # JR Assign Date (posting_start_date)
            f"<td style='padding:5px 8px;border-bottom:1px solid #e8e8e8;"
            f"font-size:12px;white-space:nowrap;text-align:center;'>{assign_date}</td>"
            # Recruiter — truncate
            f"<td style='padding:5px 8px;border-bottom:1px solid #e8e8e8;"
            f"font-size:12px;max-width:120px;overflow:hidden;"
            f"text-overflow:ellipsis;white-space:nowrap;' "
            f"title='{recruiter}'>{recruiter}</td>"
            # Status badge — compact single-letter / short label
            f"<td style='padding:5px 8px;border-bottom:1px solid #e8e8e8;"
            f"text-align:center;'>"
            f"<span style='background:{hdr_color};color:#fff;border-radius:9px;"
            f"padding:2px 9px;font-size:11px;font-weight:700;'>{badge_label}</span>"
            f"</td>"
            f"</tr>"
        )

    return f"""
<p style="margin:18px 0 5px;font-size:13px;font-weight:600;color:#333;">
  {section_title} &mdash; {len(rows)} record(s)
</p>
<table style="width:100%;border-collapse:collapse;font-family:'Segoe UI',Arial,sans-serif;
              border-radius:6px;overflow:hidden;">
  <thead>
    <tr style="background:{hdr_color};">
      <th style="padding:6px 8px;text-align:left;color:#fff;font-size:11.5px;
                 font-weight:600;white-space:nowrap;">JR No</th>
      <th style="padding:6px 8px;text-align:left;color:#fff;font-size:11.5px;
                 font-weight:600;">Job Title</th>
      <th style="padding:6px 8px;text-align:center;color:#fff;font-size:11.5px;
                 font-weight:600;white-space:nowrap;">JR Assign Date</th>
      <th style="padding:6px 8px;text-align:left;color:#fff;font-size:11.5px;
                 font-weight:600;">Recruiter</th>
      <th style="padding:6px 8px;text-align:center;color:#fff;font-size:11.5px;
                 font-weight:600;">St</th>
    </tr>
  </thead>
  <tbody>{rows_html}</tbody>
</table>"""


# ─────────────────────────────────────────────
# HTML BODY
# ─────────────────────────────────────────────
def build_html_body(summary: dict, highlights: dict) -> str:
    today_str  = datetime.now().strftime("%B %d, %Y")
    active_cnt = summary.get("active", 0)
    new_cnt    = summary.get("new_jr", 0)
    deact_cnt  = summary.get("deactivated", 0)

    new_table   = _highlights_table_html(highlights.get("new_jr", []),     "new jr")
    deact_table = _highlights_table_html(highlights.get("deactivated", []), "inactive")

    return f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="UTF-8">
  <style>
    body {{ font-family:'Segoe UI',Arial,sans-serif; background:#f4f6f9; margin:0; padding:0; }}
    .wrapper {{ max-width:640px; margin:30px auto; background:#fff; border-radius:8px;
                overflow:hidden; box-shadow:0 2px 8px rgba(0,0,0,.12); }}
    .header  {{ background:#1F4E79; padding:22px 30px; color:#fff; }}
    .header h2 {{ margin:0; font-size:18px; letter-spacing:.5px; }}
    .header p  {{ margin:4px 0 0; font-size:13px; color:#cde0f5; }}
    .body    {{ padding:28px 30px; color:#333; font-size:14.5px; line-height:1.7; }}
    .summary-cards {{ display:flex; gap:12px; margin:18px 0; }}
    .card {{ flex:1; border-radius:8px; padding:14px 16px; text-align:center; }}
    .card-label {{ font-size:11.5px; font-weight:600; text-transform:uppercase;
                   letter-spacing:.5px; margin-bottom:6px; }}
    .card-value {{ font-size:26px; font-weight:700; }}
    .card-active   {{ background:#e8f0fb; }}
    .card-active   .card-label {{ color:#1a56a0; }}
    .card-active   .card-value {{ color:#1a56a0; }}
    .card-new      {{ background:#e8f5e9; }}
    .card-new      .card-label {{ color:#1e7e34; }}
    .card-new      .card-value {{ color:#1e7e34; }}
    .card-deact    {{ background:#fdecea; }}
    .card-deact    .card-label {{ color:#b71c1c; }}
    .card-deact    .card-value {{ color:#b71c1c; }}
    .legend {{ display:flex; gap:16px; margin:18px 0; flex-wrap:wrap; }}
    .legend-item {{ display:flex; align-items:center; gap:6px; font-size:13px; color:#444; }}
    .dot {{ width:12px; height:12px; border-radius:50%; display:inline-block; }}
    .dot-new      {{ background:#00B050; }}
    .dot-active   {{ background:#0070C0; }}
    .dot-inactive {{ background:#C00000; }}
    .footer {{ background:#f0f4f8; padding:16px 30px; border-top:1px solid #dde4ee;
               font-size:12.5px; color:#666; }}
    .footer strong {{ color:#1F4E79; font-size:13.5px; }}
  </style>
</head>
<body>
<div class="wrapper">
  <div class="header">
    <h2>&#128203; Daily Job Listings Report</h2>
    <p>{today_str}</p>
  </div>
  <div class="body">
    <p>Dear Team,</p>
    <p>Please find attached the <strong>data extract</strong> as requested, along with
       today's <strong>active job listings</strong> (posting end date &gt; today).</p>

    <div class="summary-cards">
      <div class="card card-active">
        <div class="card-label">Active JRs</div>
        <div class="card-value">{active_cnt}</div>
      </div>
      <div class="card card-new">
        <div class="card-label">New JRs</div>
        <div class="card-value">{new_cnt}</div>
      </div>
      <div class="card card-deact">
        <div class="card-label">Deactivated Today</div>
        <div class="card-value">{deact_cnt}</div>
      </div>
    </div>

    {new_table}
    {deact_table}

    <p style="margin-top:20px;"><strong>Excel Status Legend:</strong></p>
    <div class="legend">
      <span class="legend-item"><span class="dot dot-new"></span> <strong>New JR</strong> &ndash; Newly added posting</span>
      <span class="legend-item"><span class="dot dot-active"></span> <strong>Active</strong> &ndash; Posting end date in future</span>
      <span class="legend-item"><span class="dot dot-inactive"></span> <strong>D</strong> &ndash; Deactivated today</span>
    </div>
    <p>For full job details and requirements, please refer to the
       <a href="https://hr-data-ui-volibits.streamlit.app/#job-requirements" style="color:#1F4E79;font-weight:600;">ATS Portal</a>.</p>
    <p>If any clarification is required, our automation desk remains at your service.</p>
  </div>
  <div class="footer">
    <strong>TalentAxis</strong><br>
    For Talent Management Automations
  </div>
</div>
</body>
</html>"""


# ─────────────────────────────────────────────
# SEND VIA MICROSOFT GRAPH
# ─────────────────────────────────────────────
def send_email():
    logging.info("=== Starting daily email job ===")

    report_window_start = get_last_successful_email_at()
    report_window_end = datetime.now()
    logging.info(
        f"Report window: {report_window_start.isoformat()} to "
        f"{report_window_end.isoformat()}"
    )

    # 1. Fetch data
    records    = fetch_active_jobs()
    highlights = fetch_highlights(report_window_start, report_window_end)
    summary    = fetch_summary_counts(highlights)
    logging.info(
        f"Records: {len(records)} | Active: {summary['active']} | "
        f"New JR: {summary['new_jr']} | Deactivated today: {summary['deactivated']}"
    )

    # 2. Excel attachment
    excel_bytes = build_excel(records)
    excel_name  = f"job_listings_{datetime.now().strftime('%Y%m%d')}.xlsx"
    excel_b64   = base64.b64encode(excel_bytes).decode("utf-8")
    logging.info(f"Excel built: {len(excel_bytes):,} bytes -> {excel_name}")

    # 3. Subject  e.g. "JR Data for BS - 08 Apr 2026"
    subject = f"JR Data for BS - {datetime.now().strftime('%d %b %Y')}"

    # 4. Graph payload
    to_recipients = [
        {"emailAddress": {"address": a.strip()}}
        for a in EMAIL_TO.split(",") if a.strip()
    ]
    cc_recipients = [
        {"emailAddress": {"address": a.strip()}}
        for a in EMAIL_CC.split(",") if a.strip()
    ]

    payload = {
        "message": {
            "subject": subject,
            "body": {"contentType": "HTML", "content": build_html_body(summary, highlights)},
            "toRecipients": to_recipients,
            **({"ccRecipients": cc_recipients} if cc_recipients else {}),
            "attachments": [{
                "@odata.type":  "#microsoft.graph.fileAttachment",
                "name":         excel_name,
                "contentType":  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "contentBytes": excel_b64,
            }],
        },
        "saveToSentItems": "true",
    }

    # 5. Send
    token   = get_graph_token()
    url     = f"https://graph.microsoft.com/v1.0/users/{EMAIL_FROM}/sendMail"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    for attempt in range(3):
        try:
            resp = requests.post(url, headers=headers, json=payload, timeout=60)
            if resp.status_code == 202:
                logging.info(f"Email sent | Subject: '{subject}'")
                logging.info(f"   To: {EMAIL_TO}" + (f" | CC: {EMAIL_CC}" if EMAIL_CC else ""))
                save_successful_email_checkpoint(report_window_end)
                clear_legacy_new_jr_status()
                break
            else:
                logging.error(f"Attempt {attempt+1}: {resp.status_code} - {resp.text}")
                if attempt < 2:
                    time.sleep(5)
        except Exception as e:
            logging.error(f"Attempt {attempt+1}: {e}")
            if attempt < 2:
                time.sleep(5)
    else:
        raise RuntimeError("All 3 Graph API send attempts failed.")

    logging.info("=== Daily email job complete ===")


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────
if __name__ == "__main__":
    send_email()
